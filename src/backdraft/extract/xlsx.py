"""Workbooks: one page per sheet, rendered as a markdown table with in-band refs.

This representation is what makes cell-level attribution work, because the
coordinates sit *in the text the model reads*:

```
## Sheet: rent-roll - Values View with cell references

| Row | A | B |
|---|---|---|
| 1 | [A1] Unit | [B1] Rent |
| 2 | [A2] 101 | [B2] 2400 |
```

A model that copies `2400` out of that table has already seen `[B2]`, so the
registry can compose `bd:model:rent-roll!B2:9e2f` for it. Nothing about the
location travels in a side channel.

The hardening here was earned on real files:
hard row/column caps, an inflated-dimension placeholder page (Excel formats empty
cells and reports a million rows), trailing empty row/column trimming, and cell
truncation. NOTE: the caps mean a snapshot can be a *partial* view of a sheet;
the page title says so, which is the honest failure the spec's "failures are
data" principle asks for.

Reader choice: openpyxl alone. python-calamine tolerates more malformed files,
but openpyxl is already a dependency (column-letter arithmetic, fixtures) and one
reader means one representation to keep golden. NOTE: if malformed workbooks show
up in practice, calamine belongs behind an optional fallback, not a second
default.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook

from .base import ExtractedPage, Extractor, ExtractionError, register
from .sheet import MAX_COLS, MAX_ROWS, bounds, format_value, partial_title, render_rows

__all__ = ["MAX_COLS", "MAX_ROWS", "XlsxExtractor", "EXTRACTOR"]

# The row/column caps, the value formatting, and the table rendering live in
# `sheet` — the representation is shared with csv and xls, and it must stay
# byte-identical (tokens are content-addressed). Kept in __all__ here for the
# callers that think of the caps as workbook policy.

_format_value = format_value
"""Compatibility alias: the formatter moved to `sheet`, unchanged."""

META_MAX_BYTES = 15_000_000
"""Styling capture opens the workbook a second time in full (non-streaming)
mode, because openpyxl's read-only mode exposes per-cell styles but not column
widths, merges, or frozen panes. Files above this size skip the styling pass
rather than risk the memory; values and anchors are unaffected."""

MAX_STYLED_CELLS = 20_000
"""Cap on styled-cell entries per sheet. Beyond it, cell styles are dropped
(widths/merges/frozen survive) and the meta says so — presentation degrades,
snapshots never do."""

_MEDIA_TYPES = frozenset({"xlsx"})
_SUFFIXES = frozenset({".xlsx", ".xlsm", ".xltx", ".xltm"})


class XlsxExtractor:
    """openpyxl workbook reader. Deterministic: values in, markdown out."""

    name = "xlsx"
    version = "2"  # v2 captures styling meta; values, text and tokens unchanged
    deterministic = True

    def can_handle(self, path: Path, media_type: str) -> bool:
        return media_type in _MEDIA_TYPES or path.suffix.lower() in _SUFFIXES

    def extract(self, path: Path, config: dict) -> Iterator[ExtractedPage]:
        """Yield one page per sheet, in workbook order.

        Sheet names are yielded as the workbook spells them; the registry
        sanitizes them to the sheetref charset when it mints tokens.

        Values stream through read-only mode (the hardened path); styling
        metadata comes from a second, size-capped full parse and rides on
        `ExtractedPage.meta`. A workbook whose styling cannot be read still
        snapshots identically — meta is display context, never identity.
        """
        try:
            workbook = load_workbook(path, data_only=True, read_only=True)
        except Exception as error:  # noqa: BLE001 - openpyxl raises broadly
            raise ExtractionError(f"could not open {path}: {error}") from error
        try:
            pages: list[ExtractedPage] = []
            bounds_by_name: dict[str, tuple[int, int]] = {}
            for number, name in enumerate(workbook.sheetnames, start=1):
                page, rows, cols = self._sheet(workbook[name], number, name)
                pages.append(page)
                if rows and cols:
                    bounds_by_name[name] = (rows, cols)
        finally:
            workbook.close()
        meta = _workbook_meta(path, bounds_by_name)
        for page in pages:
            page_meta = meta.get(page.name or "")
            if page_meta:
                page = ExtractedPage(
                    number=page.number, kind=page.kind, name=page.name,
                    text=page.text, cells=page.cells, meta=page_meta,
                )
            yield page

    def _sheet(
        self, worksheet: Any, number: int, name: str
    ) -> tuple[ExtractedPage, int, int]:
        reported_rows = worksheet.max_row or 0
        reported_cols = worksheet.max_column or 0
        if reported_rows > MAX_ROWS or reported_cols > MAX_COLS:
            return ExtractedPage(
                number=number,
                kind="sheet",
                name=name,
                text=_placeholder(name, reported_rows, reported_cols),
                cells=[],
            ), 0, 0

        grid = _read_grid(worksheet)
        rows, cols = bounds(grid)
        if not rows or not cols:
            return ExtractedPage(
                number=number,
                kind="sheet",
                name=name,
                text=partial_title(name, rows, cols, reported_rows, reported_cols),
                cells=[],
            ), 0, 0

        table, cells = render_rows(grid, rows, cols)
        title = partial_title(name, rows, cols, reported_rows, reported_cols)
        return ExtractedPage(
            number=number,
            kind="sheet",
            name=name,
            text=f"{title}\n\n{table}",
            cells=cells,
        ), rows, cols


def _read_grid(worksheet: Any) -> list[list[Any]]:
    """The sheet's values, capped. Read once: read-only sheets stream."""
    return [
        list(row)
        for row in worksheet.iter_rows(
            min_row=1, max_row=MAX_ROWS, max_col=MAX_COLS, values_only=True
        )
    ]


def _workbook_meta(
    path: Path, bounds_by_name: dict[str, tuple[int, int]]
) -> dict[str, dict]:
    """Per-sheet styling metadata, or {} when it cannot be had.

    A second, full (non-streaming) parse: read-only mode exposes per-cell
    styles but not column widths, merged ranges, or frozen panes. Bounded by
    `META_MAX_BYTES` and swallowing every failure, because styling is context
    for the artifact's sheet views — losing it must never fail an ingest.
    """
    if not bounds_by_name:
        return {}
    try:
        if path.stat().st_size > META_MAX_BYTES:
            return {}
        workbook = load_workbook(path, data_only=True, read_only=False)
    except Exception:  # noqa: BLE001 - metadata is best-effort by design
        return {}
    try:
        out: dict[str, dict] = {}
        for name, (rows, cols) in bounds_by_name.items():
            try:
                meta = _sheet_meta(workbook[name], rows, cols)
            except Exception:  # noqa: BLE001
                continue
            if meta:
                out[name] = meta
        return out
    finally:
        workbook.close()


def _sheet_meta(worksheet: Any, rows: int, cols: int) -> dict:
    """One sheet's styling: a style palette, cell -> palette refs, widths,
    merges, and the frozen pane. Only non-default styling is recorded."""
    palette: list[dict] = []
    index: dict[str, int] = {}
    cell_styles: dict[str, int] = {}
    truncated = False
    for row in worksheet.iter_rows(min_row=1, max_row=rows, max_col=cols):
        for cell in row:
            style = _cell_style(cell)
            if not style:
                continue
            if len(cell_styles) >= MAX_STYLED_CELLS:
                truncated = True
                break
            key = repr(sorted(style.items()))
            if key not in index:
                index[key] = len(palette)
                palette.append(style)
            cell_styles[cell.coordinate] = index[key]
        if truncated:
            break

    meta: dict = {}
    if cell_styles and not truncated:
        meta["palette"] = palette
        meta["cells"] = cell_styles
    elif truncated:
        meta["styles_truncated"] = True

    widths = {}
    for letter, dimension in worksheet.column_dimensions.items():
        width = getattr(dimension, "width", None)
        if getattr(dimension, "customWidth", False) and width:
            widths[letter] = round(float(width), 2)
    if widths:
        meta["widths"] = widths
    merged = [str(r) for r in getattr(worksheet.merged_cells, "ranges", [])][:200]
    if merged:
        meta["merged"] = merged
    if worksheet.freeze_panes:
        meta["frozen"] = str(worksheet.freeze_panes)
    return meta


def _cell_style(cell: Any) -> dict:
    """The non-default styling of one cell, in the artifact's compact form:
    `b` bold, `fg`/`bg` six-hex colors, `fmt` the Excel number format."""
    style: dict = {}
    font = cell.font
    if font is not None:
        if font.b:
            style["b"] = 1
        fg = _rgb(font.color)
        if fg and fg != "000000":
            style["fg"] = fg
    fill = cell.fill
    if fill is not None and getattr(fill, "patternType", None) == "solid":
        bg = _rgb(fill.fgColor)
        if bg and bg != "FFFFFF":
            style["bg"] = bg
    fmt = cell.number_format
    if fmt and fmt != "General":
        style["fmt"] = str(fmt)[:60]
    return style


def _rgb(color: Any) -> str | None:
    """A color's six-hex RGB, or None for theme/indexed/absent colors.

    openpyxl reports rgb as 8-hex ARGB; an alpha of `00` is "no color".
    """
    rgb = getattr(color, "rgb", None)
    if not isinstance(rgb, str) or len(rgb) < 6:
        return None
    if len(rgb) == 8 and rgb.upper().startswith("00"):
        return None
    return rgb[-6:].upper()


def _placeholder(name: str, rows: int, cols: int) -> str:
    """The page a sheet gets when its reported dimensions are absurd."""
    return (
        f"## Sheet: {name}\n\n"
        f"**Sheet could not be processed.**\n\n"
        f"Reported dimensions ({rows:,} rows x {cols:,} columns) exceed processing "
        f"limits ({MAX_ROWS:,} x {MAX_COLS:,}). This typically occurs when Excel "
        f"applies formatting to empty cells, inflating the apparent sheet size."
    )


EXTRACTOR: Extractor = XlsxExtractor()
register(EXTRACTOR)
