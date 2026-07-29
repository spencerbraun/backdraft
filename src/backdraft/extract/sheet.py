"""The shared tabular representation: one rendering for every sheet-shaped file.

xlsx, csv, and xls all snapshot as the same markdown values table with in-band
`[B10]` cell references, because the representation *is* the citation
mechanism: a model that copies a value out of the table has already seen its
ref, and the registry composes the cell token from it. Three extractors, one
rendering, means one shape to keep golden — a workbook saved as .xls, .xlsx,
or exported to .csv produces the same table for the same values.

Moved out of xlsx.py verbatim. Tokens are content-addressed, so the functions
here must never change what they emit for existing inputs; the xlsx golden
test pins the output character for character.
"""

from __future__ import annotations

from decimal import Decimal
from math import isfinite
from typing import Any

from openpyxl.utils import get_column_letter

from ..kernel.model import CellValue

__all__ = [
    "MAX_CELL_CHARS",
    "MAX_COLS",
    "MAX_ROWS",
    "bounds",
    "cell_text",
    "format_value",
    "has_data",
    "partial_title",
    "render_rows",
]

MAX_ROWS = 2_000
"""Hard cap on rows scanned per sheet."""

MAX_COLS = 200
"""Hard cap on columns scanned per sheet."""

MAX_CELL_CHARS = 150
"""A cell longer than this is truncated with an ellipsis."""


def has_data(value: Any) -> bool:
    """True if a cell holds data rather than formatting or emptiness."""
    return value is not None and bool(str(value).strip())


def bounds(grid: list[list[Any]]) -> tuple[int, int]:
    """The (rows, cols) actually holding data — trailing empties trimmed off."""
    rows = 0
    cols = 0
    for row_index, row in enumerate(grid, start=1):
        for col_index, value in enumerate(row, start=1):
            if has_data(value):
                rows = row_index
                cols = max(cols, col_index)
    return rows, cols


def format_value(value: Any) -> str:
    """A cell as text: integers unadorned, floats at full precision, dates ISO-8601.

    Rounding here would be a correctness bug, not a display choice. The page text
    is the snapshot and the snapshot is the receipt: a cap rate stored as
    `0.0575` and written down as `0.058` cannot be value-traced against the claim
    "5.75%" that was read off it, and the cell anchor's snippet — the thing a
    reader is shown as proof — would disagree with the workbook.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        # NOTE: before the int branch — bool is an int and `str(int(True))` is "1".
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return _float_text(value)
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def _float_text(value: float) -> str:
    """A float as the shortest text that reads back as the same float.

    `repr` gives that shortest round-trip form, which is what keeps `0.0575` from
    becoming `0.057499999999999996`. Routing it through `Decimal` only drops the
    exponent, so a small rate reads as `0.000001` rather than `1e-06` — the form
    it has in the workbook, and the form a claim would quote.
    """
    if not isfinite(value):
        return repr(value)
    if value.is_integer():
        return str(int(value))
    return format(Decimal(repr(value)), "f")


def cell_text(value: Any) -> str:
    """A formatted cell, truncated, with markdown's `|` neutralized."""
    text = format_value(value).replace("\n", " ").replace("|", "\\|")
    if len(text) > MAX_CELL_CHARS:
        text = text[: MAX_CELL_CHARS - 3] + "..."
    return text


def render_rows(
    grid: list[list[Any]], rows: int, cols: int
) -> tuple[str, list[CellValue]]:
    """The markdown table and the cell values it carries.

    Every non-empty cell gets a `[B10]` prefix in the table and a `CellValue`
    whose `value` is the cell text *as it appears in the table* — the receipt has
    to be verbatim from the snapshot, truncation and all.
    """
    header = ["Row", *(get_column_letter(col) for col in range(1, cols + 1))]
    lines = [
        "| " + " | ".join(header) + " |",
        "|" + "|".join(["---"] * len(header)) + "|",
    ]
    cells: list[CellValue] = []
    for row_index in range(1, rows + 1):
        row = grid[row_index - 1] if row_index <= len(grid) else []
        rendered = [str(row_index)]
        for col_index in range(1, cols + 1):
            value = row[col_index - 1] if col_index <= len(row) else None
            text = cell_text(value)
            if text:
                ref = f"{get_column_letter(col_index)}{row_index}"
                cells.append(CellValue(ref=ref, value=text))
                text = f"[{ref}] {text}"
            rendered.append(text)
        lines.append("| " + " | ".join(rendered) + " |")
    return "\n".join(lines), cells


def partial_title(
    name: str, rows: int, cols: int, reported_rows: int, reported_cols: int
) -> str:
    """The page's heading, which says when the view is partial."""
    note = ""
    if reported_rows > rows or reported_cols > cols:
        note = f" (showing {rows}x{cols} of {reported_rows}x{reported_cols})"
    return f"## Sheet: {name}{note} - Values View with cell references"
