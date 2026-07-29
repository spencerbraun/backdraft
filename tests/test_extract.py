"""The built-in extractors, against real files.

The xlsx test is a golden: the representation is the thing that made sub-page
attribution work, so it is pinned character for character rather than
described.
"""

from __future__ import annotations

from pathlib import Path

import pytest

from backdraft.extract import ExtractionError, base
from backdraft.extract.xlsx import MAX_COLS, MAX_ROWS

# ---- selection --------------------------------------------------------------


@pytest.mark.parametrize(
    ("filename", "media_type", "expected"),
    [
        ("notes.md", "text", "text"),
        ("notes.txt", "text", "text"),
        ("report.pdf", "pdf", "pdf-text"),
        ("model.xlsx", "xlsx", "xlsx"),
    ],
)
def test_auto_picks_the_first_extractor_that_can_handle_the_file(
    filename: str, media_type: str, expected: str
) -> None:
    assert base.select(Path(filename), media_type).name == expected


def test_auto_never_picks_the_vlm_extractor() -> None:
    assert "vlm" not in base.AUTO_ORDER


def test_an_unknown_extractor_name_is_an_extraction_error() -> None:
    with pytest.raises(ExtractionError, match="unknown extractor"):
        base.get("no-such-extractor")


def test_the_vlm_extractor_is_known_by_name_even_without_the_extra() -> None:
    """Named, so the error a user gets is about the extra, not about a typo."""
    assert "vlm" in base.names()


def test_asking_for_vlm_without_the_extra_says_it_is_unavailable() -> None:
    try:
        import openai  # noqa: F401
    except ImportError:
        with pytest.raises(ExtractionError, match="unavailable"):
            base.get("vlm")
    else:  # pragma: no cover - only when the [vlm] extra is installed
        assert base.get("vlm").deterministic is False


def test_every_built_in_reports_itself() -> None:
    # `image` is the one non-deterministic auto extractor: it is the only
    # handler for its media type, and there is no keyless floor to prefer.
    for name in base.AUTO_ORDER:
        try:
            extractor = base.get(name)
        except base.ExtractionError:
            continue  # optional extra not installed in this environment
        assert extractor.name == name
        assert extractor.version
        assert extractor.deterministic is (name != "image")


# ---- text -------------------------------------------------------------------


def test_text_is_one_page(note: Path) -> None:
    pages = list(base.get("text").extract(note, {}))
    assert len(pages) == 1
    assert pages[0].number == 1
    assert pages[0].kind == "page"
    assert pages[0].cells is None


def test_text_normalizes_line_endings(tmp_path: Path) -> None:
    path = tmp_path / "crlf.txt"
    path.write_bytes(b"one\r\ntwo\r\n")
    assert list(base.get("text").extract(path, {}))[0].text == "one\ntwo\n"


def test_text_survives_undecodable_bytes(tmp_path: Path) -> None:
    path = tmp_path / "broken.txt"
    path.write_bytes(b"before \xff\xfe after")
    assert "before" in list(base.get("text").extract(path, {}))[0].text


# ---- xlsx -------------------------------------------------------------------

GOLDEN = """\
## Sheet: Rent Roll (2025) - Values View with cell references

| Row | A | B | C |
|---|---|---|---|
| 1 | [A1] Unit | [B1] Tenant | [C1] Rent |
| 2 | [A2] 101 | [B2] Acme Corp | [C2] 2400 |
| 3 | [A3] 102 | [B3] Beta LLC | [C3] 1875.5 |"""


def test_xlsx_renders_the_golden_representation(workbook: Path) -> None:
    pages = list(base.get("xlsx").extract(workbook, {}))
    assert pages[0].text == GOLDEN


def test_xlsx_is_one_page_per_sheet(workbook: Path) -> None:
    pages = list(base.get("xlsx").extract(workbook, {}))
    assert [(page.number, page.kind, page.name) for page in pages] == [
        (1, "sheet", "Rent Roll (2025)"),
        (2, "sheet", "Summary"),
    ]


def test_xlsx_populates_cells_with_the_values_in_the_table(workbook: Path) -> None:
    page = list(base.get("xlsx").extract(workbook, {}))[0]
    assert [(cell.ref, cell.value) for cell in page.cells] == [
        ("A1", "Unit"),
        ("B1", "Tenant"),
        ("C1", "Rent"),
        ("A2", "101"),
        ("B2", "Acme Corp"),
        ("C2", "2400"),
        ("A3", "102"),
        ("B3", "Beta LLC"),
        ("C3", "1875.5"),
    ]


def test_every_cell_value_is_verbatim_in_the_page_text(workbook: Path) -> None:
    """The receipt has to quote the snapshot, so the cell must be a substring."""
    for page in base.get("xlsx").extract(workbook, {}):
        for cell in page.cells or ():
            assert f"[{cell.ref}] {cell.value}" in page.text


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        (5, "5"),
        (5.0, "5"),
        (5.25, "5.25"),
        # Full precision, not three places: the snapshot is the receipt.
        (5.123456, "5.123456"),
        (0.0575, "0.0575"),
        (0.0625, "0.0625"),
        (0.0599, "0.0599"),
        # No exponent: a small rate reads the way the workbook holds it.
        (0.000001, "0.000001"),
        (1234567890123.5, "1234567890123.5"),
        (-0.0575, "-0.0575"),
        (True, "True"),
        ("plain", "plain"),
    ],
)
def test_xlsx_formats_numbers(tmp_path: Path, value: object, expected: str) -> None:
    from openpyxl import Workbook

    book = Workbook()
    book.active.append([value])
    book.save(tmp_path / "one.xlsx")
    page = list(base.get("xlsx").extract(tmp_path / "one.xlsx", {}))[0]
    assert page.cells[0].value == expected


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        # `repr` is the shortest round-trip form, so full precision is not the
        # same as pasting in whatever binary noise the arithmetic left behind.
        (0.1 + 0.2, "0.30000000000000004"),
        (1 / 3, "0.3333333333333333"),
        (1e-7, "0.0000001"),
        (1e20, "100000000000000000000"),
        (float("nan"), "nan"),
        (float("inf"), "inf"),
    ],
)
def test_float_formatting_is_shortest_round_trip_without_an_exponent(
    value: float, expected: str
) -> None:
    """Unit-level: openpyxl normalizes some of these on write, the formatter must not."""
    from backdraft.extract.xlsx import _format_value

    assert _format_value(value) == expected


def test_a_rate_cell_is_not_rounded_in_the_snapshot(tmp_path: Path) -> None:
    """`0.0575` rounded to `0.058` makes the receipt disagree with the workbook."""
    from openpyxl import Workbook

    book = Workbook()
    book.active.append(["Going-in cap rate", 0.0575])
    book.save(tmp_path / "rates.xlsx")
    page = list(base.get("xlsx").extract(tmp_path / "rates.xlsx", {}))[0]
    assert [cell.value for cell in page.cells] == ["Going-in cap rate", "0.0575"]
    assert "[B1] 0.0575" in page.text
    assert "0.058" not in page.text


def test_xlsx_trims_trailing_empty_rows_and_columns(tmp_path: Path) -> None:
    from openpyxl import Workbook

    book = Workbook()
    sheet = book.active
    sheet["A1"] = "only"
    sheet["E9"] = None
    book.save(tmp_path / "sparse.xlsx")
    page = list(base.get("xlsx").extract(tmp_path / "sparse.xlsx", {}))[0]
    assert page.text.endswith("| Row | A |\n|---|---|\n| 1 | [A1] only |")


def test_xlsx_truncates_a_very_long_cell(tmp_path: Path) -> None:
    from openpyxl import Workbook

    book = Workbook()
    book.active["A1"] = "x" * 500
    book.save(tmp_path / "long.xlsx")
    page = list(base.get("xlsx").extract(tmp_path / "long.xlsx", {}))[0]
    assert page.cells[0].value.endswith("...")
    assert len(page.cells[0].value) == 150


def test_xlsx_escapes_the_markdown_column_separator(tmp_path: Path) -> None:
    from openpyxl import Workbook

    book = Workbook()
    book.active["A1"] = "a|b"
    book.save(tmp_path / "pipe.xlsx")
    page = list(base.get("xlsx").extract(tmp_path / "pipe.xlsx", {}))[0]
    assert page.cells[0].value == "a\\|b"


def test_an_inflated_sheet_becomes_a_placeholder_page(tmp_path: Path) -> None:
    """Excel formats empty cells and reports a million rows; say so and move on."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment

    book = Workbook()
    sheet = book.active
    sheet["A1"] = "real"
    sheet.cell(row=MAX_ROWS + 5, column=1).alignment = Alignment(horizontal="center")
    book.save(tmp_path / "inflated.xlsx")
    page = list(base.get("xlsx").extract(tmp_path / "inflated.xlsx", {}))[0]
    assert "could not be processed" in page.text
    assert page.cells == []


def test_an_empty_sheet_yields_a_page_with_no_cells(tmp_path: Path) -> None:
    from openpyxl import Workbook

    book = Workbook()
    book.save(tmp_path / "empty.xlsx")
    page = list(base.get("xlsx").extract(tmp_path / "empty.xlsx", {}))[0]
    assert page.cells == []
    assert "Sheet: Sheet" in page.text


def test_the_caps_are_the_apps_caps() -> None:
    assert (MAX_ROWS, MAX_COLS) == (2_000, 200)


def test_a_corrupt_workbook_is_an_extraction_error(tmp_path: Path) -> None:
    path = tmp_path / "not-really.xlsx"
    path.write_bytes(b"this is not a zip archive")
    with pytest.raises(ExtractionError, match="could not open"):
        list(base.get("xlsx").extract(path, {}))


# ---- pdf-text ---------------------------------------------------------------


def _make_pdf(path: Path, pages: list[list[str]]) -> Path:
    """A tiny PDF with a text layer, generated so the fixture is readable here."""
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    sheet = canvas.Canvas(str(path), pagesize=letter)
    for lines in pages:
        text = sheet.beginText(72, 720)
        for line in lines:
            text.textLine(line)
        sheet.drawText(text)
        sheet.showPage()
    sheet.save()
    return path


@pytest.fixture
def report(tmp_path: Path) -> Path:
    return _make_pdf(
        tmp_path / "t12-audit.pdf",
        [["Debt service coverage ratio: 1.42x"], ["Net operating income: $4.1 million"]],
    )


def test_pdf_text_is_one_page_per_page(report: Path) -> None:
    pages = list(base.get("pdf-text").extract(report, {}))
    assert [page.number for page in pages] == [1, 2]
    assert all(page.kind == "page" for page in pages)
    assert "1.42x" in pages[0].text
    assert "4.1 million" in pages[1].text


def test_an_image_only_pdf_names_the_vlm_extra(tmp_path: Path) -> None:
    blank = _make_pdf(tmp_path / "scan.pdf", [[], []])
    with pytest.raises(ExtractionError, match=r"backdraft\[vlm\]"):
        list(base.get("pdf-text").extract(blank, {}))


def test_a_corrupt_pdf_is_an_extraction_error(tmp_path: Path) -> None:
    path = tmp_path / "broken.pdf"
    path.write_bytes(b"%PDF-1.4 and then nothing at all")
    with pytest.raises(ExtractionError):
        list(base.get("pdf-text").extract(path, {}))


# ---- pdf-text: paragraph reconstruction -------------------------------------
#
# `extract_text()` never emits a blank line, so without this the chunker's first
# rule is dead on every PDF. These tests are about that rule firing.

_PARAGRAPHS = [
    "Total operating expenses were $1,254,800 for the trailing twelve months, "
    "or $9,803 per unit per year across the property's 128 units. The expense "
    "ratio of 46.7% of effective gross income sits at the high end of the range "
    "reported for comparable renovated assets in this submarket, and the "
    "variance is concentrated in two lines rather than spread across the "
    "statement as a whole.",
    "Real estate taxes of $412,300 represent 32.9% of total operating expenses "
    "and are the single largest line item on the statement. The 2025 "
    "reassessment raised the taxable value from $18.4 million to $24.1 million; "
    "the borrower filed an appeal in March 2025 which remains pending, and no "
    "relief has been assumed in these figures.",
    "Insurance of $122,700 reflects the renewal completed in July 2025 at a "
    "premium 19% above the expiring policy. Utilities net of reimbursement were "
    "$70,100 for the period, and the management fee of $80,500 is charged at "
    "3.0% of effective gross income under an agreement with an affiliate of the "
    "borrower, terminable on sixty days' notice.",
]


def _make_prose_pdf(path: Path, paragraphs: list[str], *, space_after: int = 10) -> Path:
    """A PDF laid out with real paragraph spacing, via reportlab's flowables.

    `space_after=0` produces the pathological case: paragraphs typeset with no
    vertical break between them, which no geometry can distinguish from a
    single block.
    """
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate

    style = ParagraphStyle(
        "Body",
        parent=getSampleStyleSheet()["BodyText"],
        fontSize=10,
        leading=14,
        spaceAfter=space_after,
    )
    SimpleDocTemplate(str(path), pagesize=LETTER).build(
        [Paragraph(text, style) for text in paragraphs]
    )
    return path


@pytest.fixture
def prose(tmp_path: Path) -> Path:
    return _make_prose_pdf(tmp_path / "prose.pdf", _PARAGRAPHS)


def test_paragraph_spacing_becomes_a_blank_line(prose: Path) -> None:
    text = list(base.get("pdf-text").extract(prose, {}))[0].text
    assert "\n\n" in text
    assert text.count("\n\n") == len(_PARAGRAPHS) - 1


def test_each_paragraph_survives_intact(prose: Path) -> None:
    """Reconstruction only inserts blank lines; it never rewrites a line."""
    text = list(base.get("pdf-text").extract(prose, {}))[0].text
    blocks = [" ".join(block.split()) for block in text.split("\n\n")]
    assert blocks == [" ".join(para.split()) for para in _PARAGRAPHS]


def test_a_page_with_paragraphs_chunks_into_several_anchors(prose: Path) -> None:
    """The point of the whole exercise: sub-page anchors on an ordinary page."""
    from backdraft.kernel.chunking import chunk

    page = list(base.get("pdf-text").extract(prose, {}))[0]
    chunks = chunk(page.text)
    assert len(chunks) == len(_PARAGRAPHS)
    assert [c.ordinal for c in chunks] == [1, 2, 3]


def test_evenly_spaced_lines_stay_one_block(tmp_path: Path) -> None:
    """No vertical break, no paragraph break — the heuristic does not invent one."""
    flat = _make_prose_pdf(tmp_path / "flat.pdf", _PARAGRAPHS, space_after=0)
    text = list(base.get("pdf-text").extract(flat, {}))[0].text
    assert "\n\n" not in text


def test_a_single_line_page_is_unchanged(tmp_path: Path) -> None:
    one = _make_pdf(tmp_path / "one.pdf", [["Debt service coverage ratio: 1.42x"]])
    assert list(base.get("pdf-text").extract(one, {}))[0].text == (
        "Debt service coverage ratio: 1.42x"
    )


def test_reconstruction_is_deterministic(prose: Path) -> None:
    """Anchors are named by ordinal; a chunker fed a wobbling page would move them."""
    runs = [list(base.get("pdf-text").extract(prose, {}))[0].text for _ in range(3)]
    assert runs[0] == runs[1] == runs[2]


# --- auto prefers the VLM for PDFs only on explicit, backdraft-scoped consent


class _StubVlm:
    name = "vlm"
    version = "stub"
    deterministic = False

    def can_handle(self, path, media_type):
        return media_type == "pdf"

    def extract(self, path, config):  # pragma: no cover - selection-only stub
        return iter(())


AMBIENT = ("OPENAI_API_KEY", "OPENROUTER_API_KEY", "ANTHROPIC_API_KEY", "OPENAI_BASE_URL")
SCOPED = ("BACKDRAFT_VLM_API_KEY", "BACKDRAFT_VLM_MODEL", "BACKDRAFT_VLM_BASE_URL", "BACKDRAFT_HOME")


def _clean(monkeypatch) -> None:
    for name in AMBIENT + SCOPED:
        monkeypatch.delenv(name, raising=False)


def test_ambient_provider_keys_are_never_consent(tmp_path, monkeypatch) -> None:
    """The incident test: generic keys in the environment must be invisible."""
    from backdraft.extract import base, vlm_settings

    _clean(monkeypatch)
    monkeypatch.setenv("OPENAI_API_KEY", "sk-ambient")
    monkeypatch.setenv("OPENROUTER_API_KEY", "sk-or-ambient")
    monkeypatch.setitem(base.EXTRACTORS, "vlm", _StubVlm())
    monkeypatch.chdir(tmp_path)
    assert not base.vlm_ready()
    pdf = tmp_path / "deck.pdf"
    pdf.write_bytes(b"%PDF-1.4")
    assert base.select(pdf, "pdf").name == "pdf-text"
    import pytest as _pytest

    from backdraft.extract.base import ExtractionError

    with _pytest.raises(ExtractionError):
        vlm_settings.client_settings({})


def test_a_scoped_env_key_is_consent(tmp_path, monkeypatch) -> None:
    from backdraft.extract import base, vlm_settings

    _clean(monkeypatch)
    monkeypatch.setenv("BACKDRAFT_VLM_API_KEY", "sk-mine")
    monkeypatch.setitem(base.EXTRACTORS, "vlm", _StubVlm())
    monkeypatch.chdir(tmp_path)
    pdf = tmp_path / "deck.pdf"
    pdf.write_bytes(b"%PDF-1.4")
    assert base.select(pdf, "pdf").name == "vlm"
    model, key, base_url = vlm_settings.client_settings({})
    assert (model, key, base_url) == (
        vlm_settings.DEFAULT_MODEL,
        "sk-mine",
        vlm_settings.OPENROUTER_BASE_URL,
    )


def test_the_env_file_is_consent(tmp_path, monkeypatch) -> None:
    from backdraft.extract import base, vlm_settings

    _clean(monkeypatch)
    monkeypatch.chdir(tmp_path)
    (tmp_path / ".backdraft").mkdir()
    (tmp_path / ".backdraft" / "env").write_text(
        "# darnell trial\nBACKDRAFT_VLM_API_KEY=sk-file\nBACKDRAFT_VLM_MODEL='my/model'\n"
    )
    monkeypatch.setitem(base.EXTRACTORS, "vlm", _StubVlm())
    assert base.vlm_ready()
    model, key, _ = vlm_settings.client_settings({})
    assert (model, key) == ("my/model", "sk-file")


def test_explicit_config_beats_every_scoped_default(tmp_path, monkeypatch) -> None:
    from backdraft.extract import vlm_settings

    _clean(monkeypatch)
    monkeypatch.setenv("BACKDRAFT_VLM_API_KEY", "sk-env")
    monkeypatch.chdir(tmp_path)
    model, key, base_url = vlm_settings.client_settings(
        {"model": "my/model", "api_key": "sk-flag", "base_url": "http://localhost:8000/v1"}
    )
    assert (model, key, base_url) == ("my/model", "sk-flag", "http://localhost:8000/v1")


def test_direct_openai_is_explicit_base_url_plus_model(tmp_path, monkeypatch) -> None:
    from backdraft.extract import vlm_settings

    _clean(monkeypatch)
    monkeypatch.setenv("BACKDRAFT_VLM_API_KEY", "sk-mine")
    monkeypatch.setenv("BACKDRAFT_VLM_BASE_URL", "https://api.openai.com/v1")
    monkeypatch.setenv("BACKDRAFT_VLM_MODEL", "gpt-4o-mini")
    monkeypatch.chdir(tmp_path)
    model, _, base_url = vlm_settings.client_settings({})
    assert (model, base_url) == ("gpt-4o-mini", "https://api.openai.com/v1")


def test_vlm_module_imports_agree_with_vlm_settings() -> None:
    """vlm.py cannot be imported in a dev env (the extra is absent), so its
    import line against vlm_settings is checked statically: every name it
    imports from `.vlm_settings` must actually exist there. This is the test
    that was missing when a deleted constant shipped as an ImportError."""
    import ast
    import pathlib

    from backdraft.extract import vlm_settings

    source = pathlib.Path(vlm_settings.__file__).with_name("vlm.py").read_text()
    tree = ast.parse(source)
    imported = [
        alias.name
        for node in ast.walk(tree)
        if isinstance(node, ast.ImportFrom) and node.module == "vlm_settings"
        for alias in node.names
    ]
    assert imported, "vlm.py no longer imports from vlm_settings?"
    missing = [name for name in imported if not hasattr(vlm_settings, name)]
    assert missing == [], missing


# --- the VLM resilience settings (stdlib, testable without the extra)


def test_run_ordered_preserves_input_order_and_reports_progress() -> None:
    import time as _time

    from backdraft.extract.vlm_settings import run_ordered

    seen: list[tuple[int, int]] = []

    def slow_square(n: int) -> int:
        _time.sleep(0.01 * (5 - n))  # later items finish first
        return n * n

    results = list(run_ordered([1, 2, 3, 4], slow_square, workers=4, progress=lambda d, t: seen.append((d, t))))
    assert results == [1, 4, 9, 16]
    assert seen == [(1, 4), (2, 4), (3, 4), (4, 4)]


def test_retry_classification_matches_production_rules() -> None:
    from backdraft.extract.vlm_settings import is_retryable

    class Http:
        def __init__(self, code):
            self.status_code = code

    class HttpError(Exception, Http):
        def __init__(self, code):
            Exception.__init__(self, f"http {code}")
            self.status_code = code

    assert is_retryable(TimeoutError())
    assert is_retryable(HttpError(429))
    assert is_retryable(HttpError(503))
    assert not is_retryable(HttpError(401))
    assert is_retryable(Exception("Rate limit exceeded"))
    assert is_retryable(Exception("upstream connect error"))
    assert is_retryable(ConnectionError("boom"))
    assert not is_retryable(Exception("invalid api key"))


def test_with_retries_backs_off_then_raises_the_last_error() -> None:
    import pytest as _pytest

    from backdraft.extract.vlm_settings import with_retries

    delays: list[float] = []
    calls = {"n": 0}

    def flaky():
        calls["n"] += 1
        raise TimeoutError(f"attempt {calls['n']}")

    with _pytest.raises(TimeoutError, match="attempt 4"):
        with_retries(flaky, attempts=4, base_delay=1.0, sleep=delays.append, jitter=lambda h: 0.0)
    assert calls["n"] == 4
    assert delays == [1.0, 2.0, 4.0]


def test_with_retries_fails_fast_on_non_retryable() -> None:
    import pytest as _pytest

    from backdraft.extract.vlm_settings import with_retries

    calls = {"n": 0}

    def broken():
        calls["n"] += 1
        raise ValueError("invalid api key")

    with _pytest.raises(ValueError):
        with_retries(broken, attempts=4, sleep=lambda s: None)
    assert calls["n"] == 1


def test_with_retries_returns_on_late_success() -> None:
    from backdraft.extract.vlm_settings import with_retries

    calls = {"n": 0}

    def eventually():
        calls["n"] += 1
        if calls["n"] < 3:
            raise TimeoutError("not yet")
        return "page text"

    assert with_retries(eventually, attempts=4, sleep=lambda s: None) == "page text"
    assert calls["n"] == 3


# ---- xlsx styling meta ------------------------------------------------------


def test_xlsx_captures_styling_meta(tmp_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    book = Workbook()
    sheet = book.active
    sheet.title = "Model"
    sheet["A1"] = "Purchase price"
    sheet["B1"] = 24850000
    sheet["A1"].font = Font(b=True)
    sheet["B1"].number_format = '"$"#,##0'
    sheet["B2"] = 0.0765
    sheet["B2"].number_format = "0.00%"
    sheet["B2"].fill = PatternFill("solid", fgColor="FFF5E6AE")
    sheet.column_dimensions["A"].width = 30
    sheet.freeze_panes = "A2"
    book.save(tmp_path / "styled.xlsx")

    page = list(base.get("xlsx").extract(tmp_path / "styled.xlsx", {}))[0]
    meta = page.meta
    assert meta is not None
    styles = {ref: meta["palette"][i] for ref, i in meta["cells"].items()}
    assert styles["A1"] == {"b": 1}
    assert styles["B1"]["fmt"] == '"$"#,##0'
    assert styles["B2"]["fmt"] == "0.00%"
    assert styles["B2"]["bg"] == "F5E6AE"
    assert meta["widths"]["A"] == 30.0
    assert meta["frozen"] == "A2"


def test_xlsx_meta_never_changes_the_snapshot(tmp_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    plain = Workbook()
    plain.active["A1"] = "x"
    plain.save(tmp_path / "plain.xlsx")
    styled = Workbook()
    styled.active["A1"] = "x"
    styled.active["A1"].font = Font(b=True)
    styled.save(tmp_path / "styled.xlsx")

    page_plain = list(base.get("xlsx").extract(tmp_path / "plain.xlsx", {}))[0]
    page_styled = list(base.get("xlsx").extract(tmp_path / "styled.xlsx", {}))[0]
    assert page_plain.text == page_styled.text
    assert [c.value for c in page_plain.cells] == [c.value for c in page_styled.cells]


# ---- csv --------------------------------------------------------------------


def test_csv_is_one_sheet_with_cell_refs(tmp_path: Path) -> None:
    path = tmp_path / "rent-roll.csv"
    path.write_text("Unit,Rent\n101,2400\n102,2450\n", encoding="utf-8")
    pages = list(base.get("csv").extract(path, {}))
    assert len(pages) == 1
    page = pages[0]
    assert page.kind == "sheet"
    assert page.name == "rent-roll"
    assert "[B2] 2400" in page.text
    refs = {c.ref: c.value for c in page.cells}
    assert refs["A1"] == "Unit" and refs["B3"] == "2450"


def test_csv_sniffs_semicolons_and_tabs(tmp_path: Path) -> None:
    semi = tmp_path / "semi.csv"
    semi.write_text("a;b\n1;2\n", encoding="utf-8")
    page = list(base.get("csv").extract(semi, {}))[0]
    assert {c.ref for c in page.cells} == {"A1", "B1", "A2", "B2"}

    tsv = tmp_path / "tabs.tsv"
    tsv.write_text("a\tb\n1\t2\n", encoding="utf-8")
    page = list(base.get("csv").extract(tsv, {}))[0]
    assert {c.ref for c in page.cells} == {"A1", "B1", "A2", "B2"}


def test_csv_tolerates_bom_and_latin1(tmp_path: Path) -> None:
    bom = tmp_path / "bom.csv"
    bom.write_bytes("﻿a,b\n1,2\n".encode("utf-8"))
    page = list(base.get("csv").extract(bom, {}))[0]
    assert {c.value for c in page.cells} == {"a", "b", "1", "2"}

    latin = tmp_path / "latin.csv"
    latin.write_bytes(b"caf\xe9,b\n")
    page = list(base.get("csv").extract(latin, {}))[0]
    assert "café" in {c.value for c in page.cells}


def test_csv_empty_file_is_an_empty_sheet(tmp_path: Path) -> None:
    path = tmp_path / "empty.csv"
    path.write_text("", encoding="utf-8")
    page = list(base.get("csv").extract(path, {}))[0]
    assert page.cells == []
    assert page.kind == "sheet"


def test_csv_reports_partial_view_when_capped(tmp_path: Path) -> None:
    from backdraft.extract.xlsx import MAX_ROWS

    path = tmp_path / "big.csv"
    path.write_text("\n".join(f"r{i},v{i}" for i in range(MAX_ROWS + 5)), encoding="utf-8")
    page = list(base.get("csv").extract(path, {}))[0]
    assert "showing" in page.text.splitlines()[0]


# ---- image ------------------------------------------------------------------


def test_image_media_type_and_selection(tmp_path: Path) -> None:
    from backdraft.registry.store import media_type_for

    assert media_type_for(Path("scan.png")) == "image"
    assert media_type_for(Path("scan.jpeg")) == "image"
    # without a key, auto must fail with the image-specific nudge, not a
    # generic no-handler error, and never fall through to the text extractor
    try:
        base.select(Path("scan.png"), "image", {})
    except base.ExtractionError as error:
        message = str(error)
        assert "vlm" in message.lower() or "key" in message.lower()
    else:  # pragma: no cover - only when a key is configured in the env
        pass


def test_image_extractor_transcribes_one_page(tmp_path: Path, monkeypatch) -> None:
    try:
        image_module = __import__("backdraft.extract.image", fromlist=["EXTRACTOR"])
    except Exception:
        import pytest

        pytest.skip("[vlm] extra not installed")
    from PIL import Image as PilImage

    path = tmp_path / "page.png"
    PilImage.new("RGB", (40, 20), "white").save(path)

    monkeypatch.setattr(
        image_module, "client_settings", lambda config: ("m", "k", "http://x")
    )
    monkeypatch.setattr(image_module, "OpenAI", lambda **kw: object())
    monkeypatch.setattr(
        image_module, "_transcribe", lambda client, model, p: "# A scanned page"
    )
    pages = list(image_module.EXTRACTOR.extract(path, {}))
    assert len(pages) == 1
    page = pages[0]
    assert page.kind == "page"
    assert page.text == "# A scanned page"
    assert page.image is not None and page.image.format == "webp"
