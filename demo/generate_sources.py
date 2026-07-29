"""Generate the demo corpus: one T12 summary PDF and one underwriting workbook.

Both files are checked in — this script exists so the corpus is reproducible and
so you can see that the numbers in it are made up on purpose. Bridgeview Commons
is not a real property and none of these figures describe one.

    uv run --group dev python demo/generate_sources.py

Deterministic: reportlab and openpyxl both stamp a creation date into the file,
so re-running changes the bytes even though the extracted text is identical. The
registry keys documents on their bytes, so a regenerated PDF ingests as a new
document. Regenerate only when you mean to.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib.enums import TA_JUSTIFY
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer

HERE = Path(__file__).resolve().parent
SOURCES = HERE / "sources"

# --------------------------------------------------------------------------- #
# The T12 summary
# --------------------------------------------------------------------------- #

# (heading, [paragraph, ...]) per page.
#
# Note on chunking: these paragraphs are typeset with real vertical space after
# each one (`spaceAfter` below), which is what the `pdf-text` extractor reads
# back off the page geometry to restore the blank lines pdfplumber's text layer
# does not emit. The chunker's blank-line rule then merges them into groups of
# roughly 200-2400 characters, so the demo shows several `pN.cM` anchors per page
# that fall on paragraph boundaries rather than mid-sentence.
T12_PAGES: list[tuple[str, list[str]]] = [
    (
        "Bridgeview Commons — Trailing Twelve Month Summary",
        [
            "Property: Bridgeview Commons, 4400 Halsted Avenue, Columbus, OH 43214. "
            "128 units across four three-story garden buildings on 6.2 acres, built "
            "in 1998 and substantially renovated between 2019 and 2021. This summary "
            "covers the trailing twelve months ended March 31, 2026, prepared from "
            "the borrower's monthly operating statements and reconciled against the "
            "property manager's general ledger. Figures are unaudited.",
            "Total effective gross income for the trailing twelve months was "
            "$2,684,400, against gross potential rent of $2,972,160. The resulting "
            "economic occupancy of 90.3% reflects an average physical occupancy of "
            "94.1% offset by concessions, bad debt, and vacancy loss. Physical "
            "occupancy was 96.4% in the most recent month and has not fallen below "
            "92.0% in any month of the period.",
            "Total operating expenses were $1,254,800, or $9,803 per unit per year. "
            "The expense ratio of 46.7% of effective gross income is at the high end "
            "of the submarket range, driven principally by real estate taxes, which "
            "were reassessed upward in January 2025 following the county's triennial "
            "revaluation. Net operating income for the trailing twelve months was "
            "$1,429,600.",
            "Real estate taxes of $412,300 represent 32.9% of total operating "
            "expenses and are the single largest line item. The 2025 reassessment "
            "raised the taxable value from $18.4 million to $24.1 million; the "
            "borrower filed an appeal in March 2025 which remains pending before the "
            "Franklin County Board of Revision. No relief has been assumed in this "
            "summary and no reserve has been established against an adverse outcome.",
            "The property carries a single mortgage of record, an agency loan "
            "originated in 2019 with an outstanding principal balance of $14,720,000 "
            "and a maturity date of December 1, 2026. Annual debt service under that "
            "loan is $974,900, implying a trailing twelve month debt service coverage "
            "ratio of 1.47x on the net operating income above. The loan is assumable "
            "subject to servicer approval and a one percent transfer fee, and the "
            "borrower has represented that no default or forbearance has occurred.",
            "Seasonality in the period followed the submarket's usual pattern. The "
            "four months from May through August 2025 produced 36.2% of the year's "
            "leasing activity and carried the highest concession load; the winter "
            "months produced the strongest net collections. Nothing in the monthly "
            "detail suggests a one-time item large enough to distort the annual "
            "figures, with the exception of the February 2026 HVAC program described "
            "on page three, which management expensed in a single month.",
            "Two items in this summary are estimates rather than recorded amounts. "
            "The allocation of the shared regional maintenance salary is made on a "
            "unit-count basis across three properties under the same management "
            "agreement, and the split of utility expense between the reimbursed and "
            "unreimbursed portions is derived from the ratio utility billing "
            "vendor's monthly file rather than from separate meters. Neither "
            "estimate is material to net operating income at the level of precision "
            "used here, but both should be confirmed in diligence.",
        ],
    ),
    (
        "Revenue Detail",
        [
            "Rental income of $2,548,900 accounts for 94.9% of effective gross "
            "income. Other income of $135,500 comprises pet rent, parking, utility "
            "reimbursement under a ratio utility billing system, and application and "
            "late fees. Utility reimbursement of $71,200 is the largest component of "
            "other income and has grown 8.4% year over year, tracking the increase "
            "in the underlying water and sewer charges rather than any change in the "
            "billing methodology.",
            "Concessions totalled $84,700 for the period, concentrated in the third "
            "quarter of 2025 when eleven units turned in a single sixty-day window "
            "following the expiry of a block of leases originally signed during "
            "lease-up of the renovated B building. Management has since staggered "
            "renewal terms and concessions in the two most recent months averaged "
            "$3,900 per month, down from a peak of $14,200 in August 2025.",
            "Bad debt of $46,900 represents 1.7% of gross potential rent. Of that "
            "figure, $31,200 relates to four units vacated without notice between "
            "May and September 2025; collection has been referred to counsel and no "
            "recovery is assumed. The remaining balance is spread across the "
            "portfolio in amounts consistent with the property's historical "
            "experience of roughly 0.6% of gross potential rent.",
            "Average in-place rent across the 128 units was $1,658 per month as of "
            "March 31, 2026. Asking rents on the most recent twelve leases signed "
            "averaged $1,742, a 5.1% premium to in-place, which supports a "
            "loss-to-lease of approximately $129,000 on an annualized basis if the "
            "current asking schedule holds through a full turn of the rent roll.",
            "Renewal behaviour has been stable. The trailing twelve month renewal "
            "rate was 58.7% against a submarket average of roughly 54%, and renewal "
            "rent increases averaged 3.4%. Turnover cost, measured as make-ready "
            "expense per move-out, averaged $1,140 and is booked within repairs and "
            "maintenance rather than capitalized, which is consistent with the "
            "borrower's stated policy but differs from the treatment used by two of "
            "the three comparable properties surveyed.",
            "Unit mix is 44 one-bedroom units, 38 two-bedroom one-bath units, 32 "
            "two-bedroom two-bath units, and 14 three-bedroom units. The "
            "two-bedroom two-bath units command the largest premium to submarket "
            "asking rents and turned least often during the period. No units were "
            "offline for renovation at any point in the trailing twelve months, and "
            "the property has no rent-restricted or income-restricted units.",
            "Collections have not deteriorated. Delinquency over thirty days stood "
            "at $18,400 as of March 31, 2026, against $22,700 twelve months earlier, "
            "and the number of residents on payment plans fell from nine to five "
            "over the same span. Two of the five current plans relate to households "
            "affected by a single employer's layoff announced in November 2025; both "
            "are current under their plans. Security deposits held total $96,000 and "
            "are maintained in a segregated account at the property's operating bank.",
        ],
    ),
    (
        "Expense Detail and Capital",
        [
            "Payroll and benefits of $286,400 cover an on-site manager, an assistant "
            "manager, two maintenance technicians, and a shared regional "
            "maintenance allocation. Headcount is unchanged over the period. "
            "Contract services of $198,100 include landscaping, snow removal, pest "
            "control, and trash; the snow removal contract was rebid in October 2025 "
            "at a rate 12% below the prior year.",
            "Repairs and maintenance of $164,200, or $1,283 per unit, is elevated "
            "relative to the $950 to $1,100 per unit typical of comparable "
            "renovated assets in the submarket. Approximately $38,000 of the "
            "variance is attributable to a single HVAC replacement program in the C "
            "building undertaken in February 2026, which the borrower expensed "
            "rather than capitalized.",
            "Insurance of $122,700 reflects the renewal completed in July 2025 at a "
            "premium 19% above the expiring policy. Utilities net of reimbursement "
            "were $70,100. Management fee of $80,500 is charged at 3.0% of effective "
            "gross income under an agreement with an affiliate of the borrower, "
            "terminable on sixty days' notice.",
            "Capital expenditures of $317,000 were incurred outside the operating "
            "statement and are not deducted in arriving at net operating income. "
            "The largest items were roof replacement on the A and B buildings "
            "($186,000) and parking lot resurfacing ($64,500). A replacement reserve "
            "of $250 per unit per year, or $32,000 annually, would be customary for "
            "an asset of this vintage and is not reflected in the figures above.",
            "Administrative and marketing expense of $54,600 covers advertising, "
            "leasing office costs, credit and background screening, and the "
            "property's share of a portfolio-level software subscription. Marketing "
            "spend rose in the third quarter of 2025 alongside the concession "
            "activity described on page two and has since returned to its prior run "
            "rate of roughly $2,800 per month.",
            "Deferred maintenance identified in the March 2026 property condition "
            "assessment totalled $410,000 over five years, of which $148,000 was "
            "flagged as immediate. The immediate items are the two remaining "
            "original boilers in the D building, the fire panel in the A building, "
            "and ADA path-of-travel work at the leasing office. None of that scope "
            "is reflected in the trailing twelve month figures and the borrower has "
            "not escrowed against it.",
            "Expense comparability against the two years prior is limited. The "
            "property changed management companies in April 2024, and the prior "
            "manager booked contract landscaping inside repairs and maintenance "
            "while the current manager books it in contract services. Restating the "
            "prior year on the current basis moves roughly $41,000 between the two "
            "lines and leaves total operating expenses unchanged; year over year "
            "movement in either line alone should not be read without that "
            "adjustment.",
        ],
    ),
]


def build_pdf(path: Path) -> None:
    """Write the T12 summary as a three-page PDF with a real text layer."""
    styles = getSampleStyleSheet()
    body = ParagraphStyle(
        "T12Body",
        parent=styles["BodyText"],
        fontName="Times-Roman",
        fontSize=10,
        leading=14,
        alignment=TA_JUSTIFY,
        spaceAfter=10,
    )
    heading = ParagraphStyle(
        "T12Heading",
        parent=styles["Heading1"],
        fontName="Times-Bold",
        fontSize=14,
        leading=18,
        spaceAfter=14,
    )
    doc = SimpleDocTemplate(
        str(path),
        pagesize=LETTER,
        leftMargin=1.0 * inch,
        rightMargin=1.0 * inch,
        topMargin=1.0 * inch,
        bottomMargin=1.0 * inch,
        title="Bridgeview Commons — T12 Summary",
        author="Backdraft demo corpus (fictional)",
    )
    flow: list[object] = []
    for index, (title, paragraphs) in enumerate(T12_PAGES):
        if index:
            flow.append(PageBreak())
        flow.append(Paragraph(title, heading))
        for text in paragraphs:
            flow.append(Paragraph(text, body))
        flow.append(Spacer(1, 6))
    doc.build(flow)


# --------------------------------------------------------------------------- #
# The underwriting model
# --------------------------------------------------------------------------- #

RENT_ROLL = [
    ("Unit Type", "Units", "Avg SF", "In-Place Rent", "Market Rent", "Occupancy"),
    ("1BR / 1BA", 44, 712, 1395, 1465, 0.955),
    ("2BR / 1BA", 38, 918, 1640, 1725, 0.947),
    ("2BR / 2BA", 32, 1042, 1835, 1930, 0.938),
    ("3BR / 2BA", 14, 1268, 2140, 2245, 0.929),
]

ASSUMPTIONS = [
    ("Assumption", "Value", "Unit", "Source"),
    ("Going-in cap rate", 0.0575, "rate", "Broker survey, Q1 2026"),
    ("Exit cap rate", 0.0625, "rate", "Going-in plus 50 bps"),
    ("Hold period", 5, "years", "Sponsor business plan"),
    ("Rent growth, year 1", 0.032, "rate", "CoStar submarket forecast"),
    ("Rent growth, years 2-5", 0.028, "rate", "CoStar submarket forecast"),
    ("Expense growth", 0.030, "rate", "Sponsor business plan"),
    ("Vacancy and credit loss", 0.070, "rate", "T12 economic occupancy"),
    ("Replacement reserve", 250, "$/unit/yr", "Lender requirement"),
    ("Purchase price", 24850000, "$", "Executed PSA"),
    ("Loan amount", 16152500, "$", "65% LTV"),
    ("Interest rate", 0.0615, "rate", "Term sheet, 2026-03-18"),
    ("Amortization", 30, "years", "Term sheet, 2026-03-18"),
    ("Year 1 NOI (underwritten)", 1487400, "$", "T12 NOI grown at 4.0%"),
    ("Year 1 DSCR", 1.34, "x", "Year 1 NOI / annual debt service"),
    ("Going-in yield on cost", 0.0599, "rate", "Year 1 NOI / purchase price"),
]


def build_xlsx(path: Path) -> None:
    """Write the two-sheet underwriting model."""
    book = Workbook()

    rent = book.active
    rent.title = "Rent Roll"
    for row in RENT_ROLL:
        rent.append(list(row))
    rent.append([])
    rent.append(
        [
            "Total / Weighted Avg",
            sum(row[1] for row in RENT_ROLL[1:]),
            None,
            None,
            None,
            None,
        ]
    )
    total_units = sum(row[1] for row in RENT_ROLL[1:])
    rent["C7"] = round(
        sum(row[1] * row[2] for row in RENT_ROLL[1:]) / total_units
    )
    rent["D7"] = round(
        sum(row[1] * row[3] for row in RENT_ROLL[1:]) / total_units
    )
    rent["E7"] = round(
        sum(row[1] * row[4] for row in RENT_ROLL[1:]) / total_units
    )
    rent["F7"] = round(
        sum(row[1] * row[5] for row in RENT_ROLL[1:]) / total_units, 3
    )
    rent.append([])
    rent.append(["Gross potential rent (annual)", 2972160, None, None, None, None])
    rent.append(["Economic occupancy (T12)", 0.903, None, None, None, None])
    rent.append(["Effective gross income", 2684400, None, None, None, None])

    assumptions = book.create_sheet("Assumptions")
    for row in ASSUMPTIONS:
        assumptions.append(list(row))

    # Styling, the way a sponsor's model actually looks: a dark header band,
    # inputs in modeling blue, currency and percent number formats, a frozen
    # header row. Values are untouched — the snapshot text and every token
    # are identical with or without this block.
    header_fill = PatternFill("solid", fgColor="FF1F4E79")
    for sheet in (rent, assumptions):
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")
        widths = {"A": 30, "B": 14, "C": 14, "D": 16, "E": 16, "F": 28}
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        sheet.freeze_panes = "A2"

    for row in range(2, 6):  # unit-mix money and occupancy columns
        for column, fmt in (("C", "#,##0"), ("D", '"$"#,##0'), ("E", '"$"#,##0')):
            rent[f"{column}{row}"].number_format = fmt
        rent[f"F{row}"].number_format = "0.0%"
    for cell in rent[7]:
        cell.font = Font(bold=True)
    rent["F7"].number_format = "0.0%"
    rent["B9"].number_format = '"$"#,##0'
    rent["B10"].number_format = "0.0%"
    rent["B11"].number_format = '"$"#,##0'

    input_font = Font(color="FF0000CC")  # modeling convention: inputs in blue
    percent_rows = (2, 3, 5, 6, 7, 8, 12, 16)
    money_rows = (9, 10, 11, 14)
    for row in range(2, 17):
        cell = assumptions[f"B{row}"]
        cell.font = input_font
        if row in percent_rows:
            cell.number_format = "0.00%"
        elif row in money_rows:
            cell.number_format = '"$"#,##0'

    book.save(path)


def main() -> None:
    SOURCES.mkdir(parents=True, exist_ok=True)
    pdf = SOURCES / "t12-summary.pdf"
    xlsx = SOURCES / "underwriting-model.xlsx"
    build_pdf(pdf)
    build_xlsx(xlsx)
    for produced in (pdf, xlsx):
        print(f"wrote {produced.relative_to(HERE.parent)} ({produced.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
